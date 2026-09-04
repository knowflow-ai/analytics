"""读 .xlsx：列出 sheet、规范化表头、推断列类型。

**不碰数据库。** 这里的每条判断错了都不会报错，只会让用户在几步之后对着一张长得不对的
表——列名不是他写的那个、编号列变成了可以求和的度量。纯逻辑才测得住。

类型按**单元格的实际类型**推断，不看字符串长得像什么。Excel 自己就带类型：日期是
datetime，数字是 int/float，`001` 存的就是文本。再去猜"看起来像数字的文本"会把编号列
判成度量，那正是 CLAUDE.md 里点名要避开的静默错答来源（`status_code`、`year`）。
"""

from __future__ import annotations

import io
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import date, datetime, time

from knowflow_analytics.errors import AnalyticsError

# 上限按"超了就说清楚"来定，不是性能调优：一次上传要整表读进内存做类型推断，
# 没有上限时一个几百兆的文件会把服务拖垮，而用户只会看到超时。
MAX_ROWS = 200_000
MAX_COLUMNS = 200


class WorkbookError(AnalyticsError):
    def __init__(self, message: str, *, code: str) -> None:
        super().__init__(message, code=code, stage="INGEST")


@dataclass(frozen=True)
class WorkbookColumn:
    """一列，以及它在表格里原来的样子。"""

    name: str
    source_name: str
    sql_type: str
    change: str = ""

    @property
    def renamed(self) -> bool:
        return bool(self.change)


@dataclass(frozen=True)
class SheetPreview:
    """一张 sheet 落库前的样子。``changes`` 是给用户过目的确认表。"""

    sheet: str
    row_count: int
    columns: tuple[WorkbookColumn, ...]
    dropped_columns: tuple[str, ...] = ()

    @property
    def changes(self) -> tuple[str, ...]:
        items = [f"{item.source_name or '(空表头)'} → {item.name}：{item.change}"
                 for item in self.columns if item.renamed]
        items.extend(f"{name}：整列为空，已忽略" for name in self.dropped_columns)
        return tuple(items)


def _load(data: bytes):
    from openpyxl import load_workbook  # 局部导入：解析器只在上传路径上需要

    try:
        return load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 openpyxl 的失败形态很多，对用户是同一件事
        raise WorkbookError(
            "这个文件读不出来，请确认是 .xlsx 格式且没有损坏。",
            code="WORKBOOK_UNREADABLE",
        ) from exc


def list_sheets(data: bytes) -> tuple[str, ...]:
    workbook = _load(data)
    try:
        names = tuple(workbook.sheetnames)
    finally:
        workbook.close()
    if not names:
        raise WorkbookError("这个文件里没有工作表。", code="WORKBOOK_EMPTY")
    return names


def _rows(data: bytes, sheet: str) -> list[tuple]:
    workbook = _load(data)
    try:
        if sheet not in workbook.sheetnames:
            raise WorkbookError(f"文件里没有名为「{sheet}」的工作表。", code="SHEET_NOT_FOUND")
        rows = []
        for index, row in enumerate(workbook[sheet].iter_rows(values_only=True)):
            if index > MAX_ROWS:
                raise WorkbookError(
                    f"这张表超过 {MAX_ROWS} 行，请先拆分或改用数据库数据源。",
                    code="SHEET_TOO_LARGE",
                )
            rows.append(row)
    finally:
        workbook.close()
    return rows


def _normalize_headers(header: tuple) -> list[tuple[str, str, str]]:
    """返回 (规范名, 原名, 改动说明)。

    自动处理三种脏表头，每一种都记下改了什么——用户要能在确认表里看见，否则他会在建模页
    对着一个自己没写过的列名。
    """

    taken: set[str] = set()
    result: list[tuple[str, str, str]] = []
    for position, raw in enumerate(header, start=1):
        source = "" if raw is None else str(raw)
        name = source.strip()
        change = ""
        if not name:
            name = f"第{position}列"
            change = "表头为空，按列位命名"
        elif name != source:
            change = "去掉了首尾空格"
        base = name
        suffix = 2
        while name.casefold() in taken:
            name = f"{base}_{suffix}"
            suffix += 1
            change = f"与前面的列重名，改为「{name}」"
        taken.add(name.casefold())
        result.append((name, source, change))
    return result


def _infer_sql_type(values: list[object]) -> str:
    present = [value for value in values if value is not None and value != ""]
    if not present:
        return "text"
    if all(isinstance(value, bool) for value in present):
        return "boolean"
    def _temporal(value: object) -> bool:
        return isinstance(value, (datetime, date)) and not isinstance(value, time)

    if all(_temporal(value) for value in present):
        # openpyxl 对纯日期单元格也返回 datetime。时间部分全为零就当日期，
        # 否则「开业日期」会变成时间戳，时间粒度上多出一层没有的精度。
        midnight = all(
            not isinstance(value, datetime) or value.time() == time(0, 0) for value in present
        )
        return "date" if midnight else "timestamp"
    numbers = [value for value in present if isinstance(value, (int, float))]
    if len(numbers) == len(present):
        if all(isinstance(value, int) for value in numbers):
            return "bigint"
        return "numeric"
    return "text"


def preview_sheet(data: bytes, sheet: str) -> SheetPreview:
    rows = _rows(data, sheet)
    if not rows:
        raise WorkbookError(f"「{sheet}」是空的。", code="SHEET_EMPTY")
    header, body = rows[0], rows[1:]
    if len(header) > MAX_COLUMNS:
        raise WorkbookError(
            f"这张表超过 {MAX_COLUMNS} 列。", code="SHEET_TOO_WIDE"
        )
    if not body:
        raise WorkbookError(f"「{sheet}」只有表头，没有数据行。", code="SHEET_HAS_NO_ROWS")

    normalized = _normalize_headers(header)
    columns: list[WorkbookColumn] = []
    dropped: list[str] = []
    for index, (name, source, change) in enumerate(normalized):
        values = [row[index] if index < len(row) else None for row in body]
        if all(value is None or value == "" for value in values):
            # 整列空：既推不出类型，建出来也只是一列 NULL。忽略它，但要说出来。
            dropped.append(source.strip() or f"第{index + 1}列")
            continue
        columns.append(
            WorkbookColumn(
                name=name, source_name=source, sql_type=_infer_sql_type(values), change=change
            )
        )
    if not columns:
        raise WorkbookError(f"「{sheet}」没有一列有数据。", code="SHEET_HAS_NO_COLUMNS")
    return SheetPreview(
        sheet=sheet,
        row_count=len(body),
        columns=tuple(columns),
        dropped_columns=tuple(dropped),
    )


def read_rows(data: bytes, preview: SheetPreview) -> Iterator[tuple]:
    """按预览确定下来的列产出数据行——预览看到什么，落库就是什么。"""

    rows = _rows(data, preview.sheet)
    header = _normalize_headers(rows[0])
    keep = [
        index
        for index, (name, _source, _change) in enumerate(header)
        if any(column.name == name for column in preview.columns)
    ]
    for row in rows[1:]:
        yield tuple(
            None if (index >= len(row) or row[index] == "") else row[index] for index in keep
        )
